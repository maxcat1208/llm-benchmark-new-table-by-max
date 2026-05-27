import os
import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from fill_new_benchmark_excel import fill_new_benchmark_excel  # noqa: E402


SOURCE_DIR = Path(os.environ["LLM_BENCH_NEW_SOURCE_DIR"])
TEMPLATE_XLSX = Path(os.environ["LLM_BENCH_NEW_TEMPLATE_XLSX"])


class FillNewBenchmarkExcelTest(unittest.TestCase):
    def test_fills_percentile_and_summary_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            target = Path(tmp) / "summary.xlsx"
            shutil.copy2(TEMPLATE_XLSX, target)

            result = fill_new_benchmark_excel(
                source_dir=SOURCE_DIR,
                target_xlsx=target,
                sheet_name="speed_benchmark",
                excel_model_name="Qwen3.5-122B-A10B",
                run_model_name="Qwen3.5-122B-A10B",
            )

            self.assertEqual(result["filled_percentile_cells"], 324)
            self.assertEqual(result["filled_summary_cells"], 64)
            self.assertEqual(result["concurrency_levels"], [8, 16, 32, 64])

            wb = load_workbook(target, data_only=True)
            ws = wb["speed_benchmark"]
            with (SOURCE_DIR / "parallel_8_number_80" / "benchmark_percentile.json").open(
                encoding="utf-8-sig"
            ) as f:
                p1 = json.load(f)[0]
            with (SOURCE_DIR / "parallel_8_number_80" / "benchmark_summary.json").open(
                encoding="utf-8-sig"
            ) as f:
                summary = json.load(f)

            self.assertEqual(ws["A38"].value, "Qwen3.5-122B-A10B")
            self.assertEqual(ws["B38"].value, 8)
            self.assertEqual(ws["C38"].value, 0.01)
            self.assertAlmostEqual(ws["D38"].value, p1["Latency (s)"], places=2)
            self.assertAlmostEqual(ws["E38"].value, p1["TTFT (ms)"], places=2)
            self.assertAlmostEqual(ws["J38"].value, p1["Output (tok/s)"], places=2)
            self.assertAlmostEqual(ws["M38"].value, summary["Test Duration (s)"], places=3)
            self.assertEqual(ws["P38"].value, summary["Total Requests"])
            self.assertEqual(ws["Q38"].value, summary["Success Requests"])
            self.assertEqual(ws["R38"].value, summary["Failed Requests"])


if __name__ == "__main__":
    unittest.main()
