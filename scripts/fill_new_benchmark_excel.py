from __future__ import annotations

import argparse
import json
import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SHORT_SHEET = "speed_benchmark"
LONG_SHEET = "speed_benchmark -long"
PARALLEL_LEVELS = [8, 16, 32, 64]
PERCENTILES = [0.01, 0.05, 0.1, 0.25, 0.5, 0.75, 0.9, 0.95, 0.99]
HEADERS = [
    "model ID",
    "parallel",
    "Percentiles",
    "Latency (s)",
    "TTFT (ms)",
    "ITL (ms)",
    "TPOT (ms)",
    "Input tokens",
    "Output tokens",
    "Output (tok/s)",
    "Total (tok/s)",
    "Decode (tok/s)",
    "Test Duration (s)",
    "Concurrency",
    "Request Rate (req/s)",
    "Total Requests",
    "Success Requests",
    "Failed Requests",
    "Req Throughput (req/s)",
    "Avg Latency (s)",
    "Avg Input Tokens",
    "Output Throughput (tok/s)",
    "Total Throughput (tok/s)",
    "TTFT (ms)",
    "TPOT (ms)",
    "ITL (ms)",
    "Avg Output Tokens",
    "Input Throughput (tok/s)",
]

PERCENTILE_COLUMNS = {
    "Latency (s)": 4,
    "TTFT (ms)": 5,
    "ITL (ms)": 6,
    "TPOT (ms)": 7,
    "Input tokens": 8,
    "Output tokens": 9,
    "Output (tok/s)": 10,
    "Total (tok/s)": 11,
    "Decode (tok/s)": 12,
}
SUMMARY_COLUMNS = {
    "Test Duration (s)": 13,
    "Concurrency": 14,
    "Request Rate (req/s)": 15,
    "Total Requests": 16,
    "Success Requests": 17,
    "Failed Requests": 18,
    "Req Throughput (req/s)": 19,
    "Avg Latency (s)": 20,
    "Avg Input Tokens": 21,
    "Output Throughput (tok/s)": 22,
    "Total Throughput (tok/s)": 23,
    "TTFT (ms)": 24,
    "TPOT (ms)": 25,
    "ITL (ms)": 26,
    "Avg Output Tokens": 27,
    "Input Throughput (tok/s)": 28,
}


def _round(value: Any) -> Any:
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    return value


def _as_percentile_label(value: Any) -> str:
    if isinstance(value, str) and value.endswith("%"):
        return value
    numeric = float(value)
    if numeric <= 1:
        numeric *= 100
    if numeric.is_integer():
        return f"{int(numeric)}%"
    return f"{numeric:g}%"


def _read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8-sig") as f:
        return json.load(f)


def _concurrency_from_dir(run_dir: Path) -> int | None:
    args_path = run_dir / "benchmark_args.json"
    if args_path.exists():
        args = _read_json(args_path)
        value = args.get("parallel")
        if value is not None:
            return int(value)
    match = re.search(r"parallel_(\d+)", run_dir.name)
    return int(match.group(1)) if match else None


def _top_left_for_cell(ws: Worksheet, row: int, col: int) -> tuple[int, int]:
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return merged_range.min_row, merged_range.min_col
    return row, col


def _merged_value(ws: Worksheet, row: int, col: int) -> Any:
    top_row, top_col = _top_left_for_cell(ws, row, col)
    return ws.cell(top_row, top_col).value


def _set_merged_value(ws: Worksheet, row: int, col: int, value: Any) -> tuple[int, int]:
    top_row, top_col = _top_left_for_cell(ws, row, col)
    ws.cell(top_row, top_col).value = value
    return top_row, top_col


def _copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    for col in range(1, len(HEADERS) + 1):
        _copy_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))


def _create_workbook(model_name: str) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in [SHORT_SHEET, LONG_SHEET]:
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        _append_model_block(ws, model_name)
    return wb


def _append_model_block(ws: Worksheet, model_name: str) -> tuple[int, int]:
    start_row = ws.max_row + 1
    total_rows = len(PARALLEL_LEVELS) * len(PERCENTILES)
    if start_row > 2:
        for offset in range(total_rows):
            template_row = 2 + (offset % total_rows)
            if template_row <= ws.max_row:
                _copy_row_style(ws, template_row, start_row + offset)

    for p_index, parallel in enumerate(PARALLEL_LEVELS):
        group_start = start_row + p_index * len(PERCENTILES)
        group_end = group_start + len(PERCENTILES) - 1
        ws.cell(group_start, 2).value = parallel
        ws.merge_cells(start_row=group_start, start_column=2, end_row=group_end, end_column=2)
        for col in range(13, len(HEADERS) + 1):
            ws.merge_cells(start_row=group_start, start_column=col, end_row=group_end, end_column=col)
        for offset, percentile in enumerate(PERCENTILES):
            ws.cell(group_start + offset, 3).value = percentile

    ws.cell(start_row, 1).value = model_name
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + total_rows - 1, end_column=1)
    return start_row, start_row + total_rows - 1


def _find_model_rows(ws: Worksheet, model_name: str) -> list[int]:
    rows = []
    for row in range(2, ws.max_row + 1):
        if _merged_value(ws, row, 1) == model_name:
            rows.append(row)
    return rows


def _load_or_create_workbook(target_xlsx: Path, model_name: str) -> Workbook:
    if target_xlsx.exists():
        return load_workbook(target_xlsx)
    return _create_workbook(model_name)


def _collect_evalscope_outputs(source_dir: Path) -> dict[int, dict[str, Any]]:
    outputs: dict[int, dict[str, Any]] = {}
    for run_dir in sorted(source_dir.glob("parallel_*_number_*")):
        concurrency = _concurrency_from_dir(run_dir)
        if concurrency is None:
            continue
        percentile_path = run_dir / "benchmark_percentile.json"
        summary_path = run_dir / "benchmark_summary.json"
        outputs[concurrency] = {
            "percentile": _read_json(percentile_path) if percentile_path.exists() else [],
            "summary": _read_json(summary_path) if summary_path.exists() else {},
        }
    return outputs


def _clear_model_rows(ws: Worksheet, model_rows: list[int]) -> None:
    for row in model_rows:
        for col in range(4, len(HEADERS) + 1):
            top_left = _top_left_for_cell(ws, row, col)
            if top_left == (row, col):
                ws.cell(row, col).value = None


def fill_new_benchmark_excel(
    source_dir: str | Path,
    target_xlsx: str | Path,
    sheet_name: str,
    excel_model_name: str | None = None,
    run_model_name: str | None = None,
) -> dict[str, Any]:
    source_dir = Path(source_dir)
    target_xlsx = Path(target_xlsx)
    run_model_name = run_model_name or source_dir.name
    excel_model_name = excel_model_name or run_model_name
    outputs = _collect_evalscope_outputs(source_dir)

    wb = _load_or_create_workbook(target_xlsx, excel_model_name)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        _append_model_block(ws, excel_model_name)
    ws = wb[sheet_name]

    model_rows = _find_model_rows(ws, excel_model_name)
    if not model_rows:
        start, end = _append_model_block(ws, excel_model_name)
        model_rows = list(range(start, end + 1))
    _clear_model_rows(ws, model_rows)

    row_by_key: dict[tuple[int, str], int] = {}
    group_top_by_concurrency: dict[int, int] = {}
    for row in model_rows:
        concurrency = int(_merged_value(ws, row, 2))
        percentile_label = _as_percentile_label(ws.cell(row, 3).value)
        row_by_key[(concurrency, percentile_label)] = row
        group_top_by_concurrency.setdefault(concurrency, _top_left_for_cell(ws, row, 2)[0])

    filled_percentile_cells = 0
    filled_summary_cells: set[tuple[int, int]] = set()
    missing_concurrency = []

    for concurrency, data in sorted(outputs.items()):
        if concurrency not in group_top_by_concurrency:
            missing_concurrency.append(concurrency)
            continue

        for percentile_row in data.get("percentile") or []:
            percentile_label = _as_percentile_label(percentile_row.get("Percentiles"))
            row = row_by_key.get((concurrency, percentile_label))
            if row is None:
                continue
            for key, col in PERCENTILE_COLUMNS.items():
                value = percentile_row.get(key)
                if value is not None:
                    ws.cell(row, col).value = _round(value)
                    filled_percentile_cells += 1

        summary = data.get("summary") or {}
        group_row = group_top_by_concurrency[concurrency]
        for key, col in SUMMARY_COLUMNS.items():
            value = summary.get(key)
            if value is not None:
                written_cell = _set_merged_value(ws, group_row, col, _round(value))
                filled_summary_cells.add(written_cell)

    target_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_xlsx)
    return {
        "source_dir": str(source_dir),
        "target_xlsx": str(target_xlsx),
        "sheet_name": sheet_name,
        "excel_model_name": excel_model_name,
        "run_model_name": run_model_name,
        "concurrency_levels": sorted(outputs),
        "filled_percentile_cells": filled_percentile_cells,
        "filled_summary_cells": len(filled_summary_cells),
        "missing_concurrency": missing_concurrency,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", required=True)
    parser.add_argument("--target-xlsx", default="模型测试汇总表-new.xlsx")
    parser.add_argument("--sheet-name", required=True, choices=[SHORT_SHEET, LONG_SHEET])
    parser.add_argument("--excel-model-name")
    parser.add_argument("--run-model-name")
    args = parser.parse_args()
    result = fill_new_benchmark_excel(
        source_dir=args.source_dir,
        target_xlsx=args.target_xlsx,
        sheet_name=args.sheet_name,
        excel_model_name=args.excel_model_name,
        run_model_name=args.run_model_name,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
